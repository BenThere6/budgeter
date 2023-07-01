import imaplib
import shutil
import email
from email.header import decode_header
from sqlite3 import Date
import os
from datetime import datetime
from unittest import skip
from openpyxl import load_workbook
import warnings
from os.path import exists
import time
import sys
import socket
import smtplib
from email.message import EmailMessage
import win32com.client as win32
import traceback

me = '' '''redacted'''
us = ['' '''redacted''','' '''redacted''']
her = '' '''redacted'''
current_time = datetime.now().strftime("%I:%M %p")
if os.path.exists('C:/Users/benja'):
    user = 'benja'
elif os.path.exists('C:/Users/Ben'):
    user = 'Ben'

def main():
    deleteEmail = True
    success = False
    test_internet_connection()
    wb, success = is_workbook_open(success)
    username = "" '''redacted'''
    password = "" '''redacted'''
    no_emails = get_emails(username,password,user,deleteEmail)
    more_than_one = scrape_html()
    short_details, short_amount, exp, success = to_excel(wb, success)
    print_final_message(more_than_one,short_details,short_amount,exp)
    remove_files()
    if get_part_of_day(datetime.now().hour) != 'morning':
        try:
            os.remove('C:\\Users\\'+user+'\\OneDrive\VS Code\\part.txt')
        except:
            pass
    close()

def get_part_of_day(h):
    return (
        "morning"
        if 5 <= h <= 11
        else "afternoon"
        if 12 <= h <= 16
        else "evening"
        if 17 <= h <= 22
        else "night"
    )

def close():
    if get_part_of_day(datetime.now().hour) != 'morning':
        try:
            os.remove('C:\\Users\\'+user+'\\OneDrive\VS Code\\part.txt')
        except:
            pass
    print("\nProgram will close in 15 seconds")
    time.sleep(15)
    sys.exit()

def test_internet_connection():
    repeat = True
    while repeat:
        try:
            socket.create_connection(('Google.com',80))
            repeat = False
        except:
            print("You are not connected to the internet\n")
            # if remote:
            with open("no_internet.txt",'a+') as f:
                time = datetime.now()
                f.write(f"Tried to run program, but didn't have internet - {time}")
            sys.exit()

def is_workbook_open(success):
    repeat = True
    while repeat:
        try:
            path = 'C:/Users/'+user+'/OneDrive/Money Manager_no_macro.xlsx'
            wb = load_workbook(path)
            wb.close()
            repeat = False
        except:
            message = "\nSpreadsheet is open, unable to run the program at this time."
            subject = 'ERROR '+current_time
            text_updates(subject, message, me)
            close()
    return wb, success

def delete_unrelated_emails(imap):
    try:
        imap.select("INBOX")
        status, messages = imap.search(None, 'SUBJECT "Your Mountain America statement is available"')
        messages = messages[0].split(b' ')
        for mail in messages:
            _, msg = imap.fetch(mail, "(RFC822)")
            imap.store(mail, "+FLAGS", "\\DELETED")
            print("Unrelated email deleted")
    except:
        pass
    try:
        imap.select("INBOX")
        status, messages = imap.search(None, 'FROM "donotreply@mcau.com"')
        messages = messages[0].split(b' ')
        for mail in messages:
            _, msg = imap.fetch(mail, "(RFC822)")
            imap.store(mail, "+FLAGS", "\\DELETED")
            print("Unrelated email deleted")
    except:
        pass
    try:
        imap.select("INBOX")
        status, messages = imap.search(None, '#Secure# Message from Mountain America Credit Union 2207120286C"')
        messages = messages[0].split(b' ')
        for mail in messages:
            _, msg = imap.fetch(mail, "(RFC822)")
            imap.store(mail, "+FLAGS", "\\DELETED")
            print("Unrelated email deleted")
    except:
        pass
    try:
        imap.select("INBOX")
        status, messages = imap.search(None, 'SUBJECT "Your Mountain America Visa statement is available"')
        messages = messages[0].split(b' ')
        for mail in messages:
            _, msg = imap.fetch(mail, "(RFC822)")
            imap.store(mail, "+FLAGS", "\\DELETED")
            print("Unrelated email deleted")
    except:
        pass
    try:
        imap.select("INBOX")
        status, messages = imap.search(None, 'SUBJECT "New Remote Deposit Alert from Mountain America Credit Union"')
        messages = messages[0].split(b' ')
        for mail in messages:
            _, msg = imap.fetch(mail, "(RFC822)")
            imap.store(mail, "+FLAGS", "\\DELETED")
            print("Unrelated email deleted")
    except:
        pass
    try:
        imap.select("INBOX")
        status, messages = imap.search(None, 'SUBJECT "Your feedback is vital to helping us improve"')
        messages = messages[0].split(b' ')
        for mail in messages:
            _, msg = imap.fetch(mail, "(RFC822)")
            imap.store(mail, "+FLAGS", "\\DELETED")
            print("Unrelated email deleted")
    except:
        pass

def clean(text):
    return "".join(c if c.isalnumolder() else "_" for c in text)

def get_emails(username,password,user,deleteEmail):
    # create an IMAP4 class with SSL 
    imap = imaplib.IMAP4_SSL("imap.gmail.com")
    # authenticate
    imap.login(username, password)

    # Delete unrelated emails
    delete_unrelated_emails(imap)

    imap.select("INBOX")
    status, messages = imap.search(None, 'SUBJECT "Transaction Alert from Mountain America Credit Union"')

    # total number of emails
    messages = len(messages[0])
    if messages < 1:
        print()
        print(datetime.now().strftime("%m/%d/%Y  %I:%M %p"))
        print("Did not find any emails\n")
        no_emails = True
        # if remote:
        subject = datetime.now().strftime("%m/%d/%Y  %I:%M %p")
        message = 'The program ran successfully, no transactions were found.'
        path = 'C:\\Users\\'+user+'\\OneDrive\\VS Code\\Program Files\\program_log.txt'
        with open(path,'a+') as f:
            f.write(subject+': '+message+'\n')
        print("Program log updated")
        # update_email = me
        # text_updates(subject, message, update_email)
        # print(f'Text message sent to {update_email}')
        if get_part_of_day(datetime.now().hour) == 'morning':
            if not os.path.exists('C:\\Users\\'+user+'\\OneDrive\\VS Code\\part.txt'):
                f = open('C:\\Users\\'+user+'\\OneDrive\\VS Code\\part.txt','a+')
                from_excel(True,False)
        close()
    else:
        no_emails = False

    for i in range(messages, 0, -1):
        # fetch the email message by ID
        res, msg = imap.fetch(str(i), "(RFC822)")
        for response in msg:
            if isinstance(response, tuple):
                # parse a bytes email into a message object
                msg = email.message_from_bytes(response[1])
                # decode the email subject
                subject, encoding = decode_header(msg["Subject"])[0]

                if isinstance(subject, bytes):
                    # if it's a bytes, decode to str
                    subject = subject.decode(encoding)
                # decode email sender
                From, encoding = decode_header(msg.get("From"))[0]

                if isinstance(From, bytes):
                    From = From.decode(encoding)
                Date, encoding = decode_header(msg.get("Date"))[0]

                if isinstance(Date, bytes):
                    Date = Date.decode(encoding)
                print("Date:", Date)
                print("Subject:", subject)
                print("From:", From)

                # if the email message is multipart
                if msg.is_multipart():
                    # iterate over email parts
                    for part in msg.walk():
                        # extract content type of email
                        content_type = part.get_content_type()
                        content_disposition = str(part.get("Content-Disposition"))
                        try:
                            # get the email body
                            body = part.get_payload(decode=True).decode()
                        except:
                            pass
                        if content_type == "text/plain" and "attachment" not in content_disposition:
                            # print text/plain emails and skip attachments
                            print(body)
                        elif "attachment" in content_disposition:
                            # download attachment
                            filename = part.get_filename()
                        
                            if filename:
                                folder_name = clean(subject)
                                if not os.path.isdir(folder_name):
                                    # make a folder for this email (named after the subject)
                                    os.mkdir(folder_name)
                                filepath = os.path.join(folder_name, filename)
                                # download attachment and save it
                                open(filepath, "wb").write(part.get_payload(decode=True))
                            
                else:
                    # extract content type of email
                    content_type = msg.get_content_type()
                    # get the email body
                    body = msg.get_payload(decode=True).decode()
                    if content_type == "text/plain":
                        # print only text email parts
                        print(body)
                if content_type == "text/html":
                    currentDate = datetime.now().strftime("%Y_%m_%d-%I;%M;%S_%p")
                    filename = currentDate + '.html'
                    with_new_line = filename + '\n'

                    #filepath = os.path.join(folder_path, filename)
                    path = "C:/Users/"+user+"/OneDrive/VS Code/Python Files/Auto_Email/transactions_file"
                    f = open(path,'a')
                    f.write(with_new_line)
                    f.close()
                        
                    # write the file
                    write_file_location = 'C:\\Users\\' + user + '\\OneDrive\\VS Code\\Python Files\\' +filename
                    open(write_file_location, "w").write(body)

                    time.sleep(1)
                print("="*100)

    if deleteEmail:
        imap.select("INBOX")
        status, messages = imap.search(None, 'SUBJECT "Transaction Alert from Mountain America Credit Union"')
        messages = messages[0].split(b' ')
        for mail in messages:
            _, msg = imap.fetch(mail, "(RFC822)")
            imap.store(mail, "+FLAGS", "\\DELETED")
        print("Email deleted")

    # Close the connection and logout
    imap.close()
    imap.logout()

    return no_emails

def scrape_html():
    more_than_one = False
    file_list = []
    comp_path = 'C:/Users/'+user+'/OneDrive/VS Code/Python Files/Auto_Email/transactions_file'
    with open(comp_path,'r+') as f:
        for line1 in f:
            html_str = ''
            strip_string_path = 'C:\\Users\\' + user + '\\OneDrive\\VS Code\\Python Files\\' + line1.rstrip('\n')
            with open(strip_string_path,'r+') as f:
                for line in f:
                    html_str += line

            new1 = html_str.split('Hello  redacted')  
            new2 = new1[1].split('ght: 150%; margin-top: 0; margin-right: 0; margin-left: 0; margin-bottom: 0; padding-top: 0; padding-right: 0; padding-bottom: 0; padding-left: 0;" />')
            new3 = new2[0].split('<td colspan="3"><hr style="height: 1px; margin: 0; border-style: none; color: #EEEEEE; background-color: #EEEEEE;" /></td>')
            
            del new3[-1]

            detes = []
            detes2 = []
            amounts = []
            amounts2 = []

            for item in new3:
                new_dete = item.split('line-height: 1.4; min-height: 43px;">                          ')
                detes.append(new_dete[1])

            for item in detes:
                new_dete = item.split('</td>                      <td class="amount trans-amount"')
                new_new = new_dete[0]
                detes2.append(new_new)

            for item in new3:
                new_amount = item.split('px;text-align: right;">                          ($')
                amounts.append(new_amount[1])

            for item in amounts:
                new_amount = item.split(')                      </td>                  </tr>')
                amounts2.append(new_amount[0])

            # Save transactions to a file as receipt
            time = datetime.now().strftime("%m/%d/%Y  %I:%M %p")
            path = 'C:/Users/'+user+'/OneDrive/VS Code/Program Files/Transaction Receipts.txt'
            holder = 0
            for line in detes2:
                real_details = line
                final_amt_transferred = amounts2[holder]
                with open(path,'a') as receipt:
                    with open('C:\\Users\\'+user+'\\OneDrive\VS Code\\Python Files\\transaction_details.txt','a') as f:
                        to_write = time + ', ' + real_details + ', ' + final_amt_transferred + '\n'
                        receipt.write(to_write)
                        if len(to_write) < 200:
                            f.write(to_write)
                holder += 1
            
            
            # Determine if the email contains one transaction, or two+ transactions successionally
            length = len(html_str)
            if length > 11500:
                more_than_one = True
        print("Receipt updated")
        file_list.append(line.rstrip('\n'))
    return more_than_one

def receive_reply(filter_old):
    email_user = '' '''redacted'''
    email_pass = '' '''redacted'''
    path = 'C:\\Users\\'+user+'\\OneDrive\\VS Code\\Program Files\\Transaction Organization\\replies\\text_0.txt'
    if os.path.exists(path):
        os.remove(path)

    received = False
    count = 0
    while received == False:
        count += 1
        if count == 2 and filter_old == False:
            print("\nWaiting for text...")
        mail = imaplib.IMAP4_SSL('imap.gmail.com')
        mail.login(email_user, email_pass)
        mail.select('Inbox')

        type, data = mail.search(None, 'UNSEEN')
        files = []
        if len(data[0].split()) == 0 and filter_old == True:
            received = True
        for num in data[0].split():
            typ, data = mail.fetch(num, '(RFC822)' )
            raw_email = data[0][1]
            raw_email_string = raw_email.decode('utf-8')
            email_message = email.message_from_string(raw_email_string)
            if not filter_old:
                # downloading attachments
                for part in email_message.walk():
                    if part.get_content_maintype() == 'multipart':
                        continue
                    if part.get('Content-Disposition') is None:
                        continue
                    fileName = part.get_filename()
                    if fileName not in files or fileName in files:
                        files.append(fileName)
                        if bool(fileName):
                            filePath = os.path.join('C:\\Users\\'+user+'\\OneDrive\\VS Code\\Program Files\\Transaction Organization\\attachments\\', fileName)
                            filePath2 = os.path.join('C:\\Users\\'+user+'\\OneDrive\\VS Code\\Program Files\\Transaction Organization\\replies\\', fileName)
                            if not os.path.isfile(filePath) :
                                fp = open(filePath, 'wb')
                                fp.write(part.get_payload(decode=True))
                                fp.close()
                            print(f'Downloaded "{fileName}".')
                            if fileName == 'text_0.txt':
                                shutil.move(filePath, filePath2)
                                received = True
            else:
                received = True

def sep_reply():
    with open('C:\\Users\\'+user+'\\OneDrive\\VS Code\\Program Files\\Transaction Organization\\replies\\text_0.txt','r') as f:
        lines = []
        for line in f:
            lines.append(line)
        short_filter = lines[0].strip()
        print(short_filter)
        try:
            category_filter = int(lines[1].strip())
        except:
            category_filter = 7
        print(category_filter)
    os.remove('C:\\Users\\'+user+'\\OneDrive\\VS Code\\Program Files\\Transaction Organization\\replies\\text_0.txt')
    return short_filter, category_filter

def to_excel(wb, success):
    warnings.filterwarnings('ignore',category=UserWarning,module='openpyxl')
    short_details = []
    short_amount = []
    exp = False
    comp_path = 'C:/Users/'+user+'/OneDrive/VS Code/Program Files/Transaction Organization/'
    repeat = True
    while repeat:
        if os.path.exists(comp_path):
            repeat = False
        else:
            text_updates('ERROR '+current_time,'Transaction Organization folder is not in the correct location. Must be replaced before program can run.',me)
            if get_part_of_day(datetime.now().hour) == 'morning':
                if not os.path.exists('C:\\Users\\'+user+'\\OneDrive\\VS Code\\part.txt'):
                    f = open('C:\\Users\\'+user+'\\OneDrive\\VS Code\\part.txt','a+')
                    from_excel(True, success)
            close()
        
    expenses = []
    with open(comp_path + 'expenses.txt', 'r+') as e:
        for line in e:
            if line != '' and line != '\n':
                lyne = line.strip()
                expenses.append(lyne)

    internal = []
    with open(comp_path + 'internal.txt', 'r+') as e:
        for line in e:
            if line != '' and line != '\n':
                lyne = line.strip()
                internal.append(lyne)
    
    other = []
    with open(comp_path + 'other.txt', 'r+') as e:
        for line in e:
            if line != '' and line != '\n':
                lyne = line.strip('\n')
                other.append(lyne)

    food = []
    with open(comp_path + 'food.txt', 'r+') as e:
        for line in e:
            if line != '' and line != '\n':
                lyne = line.strip('\n')
                food.append(lyne)

    gas = []
    with open(comp_path + 'gas.txt', 'r+') as e:
        for line in e:
            if line != '' and line != '\n':
                lyne = line.strip('\n')
                gas.append(lyne)

    shopping = []
    with open(comp_path + 'shopping.txt', 'r+') as e:
        for line in e:
            if line != '' and line != '\n':
                lyne = line.strip('\n')
                shopping.append(lyne)

    f = open('C:\\Users\\'+user+'\\OneDrive\VS Code\\Python Files\\transaction_details.txt','r+')
    for line in f:
        line_list = line.split(',')
        repeat = True
        ws = wb['Transactions']
        while repeat:
            unfiltered = True
            for place in food:
                if place in line:
                    trans_type = 'Food'
                    unfiltered = False
                    repeat = False
                    short_details.append(place)
            if repeat:
                for place in gas:
                    if place in line:
                        trans_type = 'Gas'
                        unfiltered = False
                        repeat = False
                        short_details.append(place)
            if repeat:        
                for place in shopping:
                    if place in line:
                        trans_type = 'Shopping'
                        unfiltered = False
                        repeat = False
                        short_details.append(place)
            if repeat:
                for place in expenses:
                    if place in line:
                        trans_type = 'Expenses'
                        unfiltered = False
                        repeat = False
                        short_details.append(place)
                        exp = True
            if repeat:
                for place in internal:
                    if place in line:
                        trans_type = 'Internal'
                        unfiltered = False
                        repeat = False
                        short_details.append(place)
            if repeat:
                for place in other:
                    if place in line:
                        trans_type = 'Other'
                        unfiltered = False
                        repeat = False
                        short_details.append(place)
            if unfiltered == True:
                print("-"*35)
                print("These details are unfiltered:\n")
                unf_detes = line_list[1]+"\n"
                print(line_list[1]+"\n")

                message = 'This transaction is unfiltered: \n'+unf_detes+'\n1. Shopping, 2. Food, 3. Gas, 4. Other, 5. Internal, 6. Expenses\nPlease reply with the short details on the first line, and the category number on the second line'
                receive_reply(True)
                text_updates('Unfiltered Transaction '+current_time,message,me)
                receive_reply(False)
                short_filter, category_filter = sep_reply()
                while category_filter != 1 and category_filter != 2 and category_filter != 3 and category_filter != 4 and category_filter != 5 and category_filter != 6:
                    text_updates('ERROR '+current_time,'\nThe catergory filter # you entered is invalid.',me)
                    # text_updates('Unfiltered Transaction',message,me)
                    receive_reply(False)
                    short_filter, category_filter = sep_reply()
                print('The function worked, here are the filters:')
                print(f'short:{short_filter} category: {category_filter}')

                if category_filter == 1:
                    shopping.append(short_filter)
                    path_code = "shopping.txt"
                elif category_filter == 2:
                    food.append(short_filter)
                    path_code = "food.txt"
                elif category_filter == 3:
                    gas.append(short_filter)
                    path_code = "gas.txt"
                elif category_filter == 4:
                    other.append(short_filter)
                    path_code = "other.txt"
                elif category_filter == 5:
                    internal.append(short_filter)
                    path_code = "internal.txt"
                elif category_filter == 6:
                    expenses.append(short_filter)
                    path_code = "expenses.txt"
                else:
                    expenses.append(short_filter)
                    path_code = "mistakes.txt"

                filter_path = 'C:\\Users\\'+user+'\\OneDrive\\VS Code\\Program Files\\Transaction Organization\\'+path_code
                with open(filter_path,'a') as qt:
                    qt.write(f"\n{short_filter}")
                    print("New filter added successfully")
            else:
                repeat = False
            if repeat == False:
                temp_list_1 = line_list[0]
                temp_list_2 = line_list[1]
                temp_list_3 = line_list[2].strip()
                if temp_list_3[-1] == '.':
                    temp_list_3 = temp_list_3.rstrip(temp_list_3[-1])
                if temp_list_3[0] == '$':
                    temp_list_3 = temp_list_3.replace('$','')
                try:
                    temp_list_3 = float(temp_list_3)
                except:
                    temp_list_3 = float(line_list[3].strip())
                temp_list = [temp_list_1,temp_list_2,temp_list_3]
                sepfail = False
                if not sepfail:
                    short_amount.append(temp_list_3)
                ready = True
            else:
                ready = False
        if ready:
            temp_list.insert(0,trans_type)
            ws.append(temp_list)

    ws = wb['Minutia']

    f.close()
    path = 'C:/Users/'+user+'/OneDrive/Money Manager_no_macro.xlsx'
    wb.save(path)

    
    print("Workbook updated")
    success = True
    from_excel(False)
    return short_details, short_amount, exp, success

def from_excel(morning, success=True):
    try:
        excel = win32.gencache.EnsureDispatch('Excel.Application')
    except:
        path = 'C:\\Users\\'+user+'\\AppData\\Local\\Temp\\gen_py'
        shutil.rmtree(path)
        time.sleep(5)
        excel = win32.gencache.EnsureDispatch('Excel.Application')
    wb = excel.Workbooks.Open('C:/Users/'+user+'/OneDrive/Money Manager_no_macro.xlsx')
    excel.Visible = False
    wb.Save()
    # wb.close()
    excel.Application.Quit()
    print('Waiting 20 seconds for excel to load correctly')
    time.sleep(20)
    path = 'C:/Users/'+user+'/OneDrive/Money Manager_no_macro.xlsx'
    wb = load_workbook(path, data_only=True)
    ws = wb['Minutia']
    target_perc = str(int(ws['I6'].value*100))
    food_perc = str(int(ws['H13'].value*100))
    shopping_perc = str(int(ws['I13'].value*100))
    gas_perc = str(int(ws['J13'].value*100))
    other_perc = str(int(ws['K13'].value*100))
    current_perc = str(int(ws['J6'].value*100))

    food_target = str(int(ws['G48'].value))
    shopping_target = str(int(ws['H48'].value))
    gas_target = str(int(ws['I48'].value))
    other_target = str(int(ws['J48'].value))

    if morning and not success:
        print("morrrrning")
        subject = 'Morning Review '+current_time
        message = '\nWe are '+target_perc+'% through the month. '+current_perc+'% through the budget.\n'+food_perc+'% Food         $'+food_target+'\n'+shopping_perc+'% Shopping     $'+shopping_target+'\n'+gas_perc+'% Gas           $'+gas_target+'\n'+other_perc+'% Other            $'+other_target+'\n'
    else:
        print('Not morning')
        subject = 'Update at '+current_time
        message = '\nWe are '+target_perc+'% through the month. '+current_perc+'% through the budget.\n'+food_perc+'% Food         $'+food_target+'\n'+shopping_perc+'% Shopping     $'+shopping_target+'\n'+gas_perc+'% Gas           $'+gas_target+'\n'+other_perc+'% Other            $'+other_target+'\n'
        # print(len(message))
    
    text_updates(subject,message,us)
    print("Text update sent")
    # print('Maybe waiting 30 seconds will help?')
    # time.sleep(30)
    wb.close()
    
def print_final_message(more_than_one,short_details,short_amount,exp):
    if len(short_details)>0:
        print("-"*35)
    q = 0
    message = ''
    for_message = []
    for i in short_details:
        print(i," --> $",short_amount[q])
        for_message.append('\n'+i+' --> $'+str(short_amount[q]))
        q += 1
    if not more_than_one:
        more_than_one = ''
    if not exp:
        exp = ''
    print("-"*35)
    print(f"Multiple:   {more_than_one}")
    print(f"Expenses:   {exp}")
    print("-"*35)

    for item in for_message:
        message = message + item

    print('Waiting 1 minute before sending transactions text')
    time.sleep(60)
    text_updates('Transactions '+str(current_time),message,me)
    print('Transactions text sent')

def remove_files():
    this_path = 'C:/Users/' + user + '/OneDrive/VS Code/Python Files/Auto_Email/transactions_file'
    with open(this_path) as f:
        for filename in f:
            filename = filename.rstrip('\n')
            path_to_filename = 'C:\\Users\\' + user + '\\OneDrive\\VS Code\\Python Files\\' + filename
            os.remove(path_to_filename)

    path = 'C:/Users/'+user+'/OneDrive/VS Code/Python Files/Auto_Email/transactions_file'
    try:
        os.remove(path)
        print("Filenames removed")
    except:
        print('Unable to remove transactions file')
        text_updates('ERROR '+current_time,'Unable to remove transactions file',me)
    try:
        os.remove('C:\\Users\\'+user+'\\OneDrive\VS Code\\Python Files\\transaction_details.txt')
        print("Details removed")
    except:
        print('Unable to remove transaction details')
        text_updates('ERROR '+current_time,'Unable to remove transaction details',me)

    return 

def text_updates(subject, message, update_email):
    msg = EmailMessage()
    msg.set_content(message)

    msg['Subject'] = subject
    msg['From'] = "" '''redacted'''
    msg['To'] = update_email

    server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
    server.login("" '''redacted''', "" '''redacted''')
    server.send_message(msg)
    server.quit()

main()
